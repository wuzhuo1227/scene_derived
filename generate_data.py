# coding=utf-8
import numpy as np
import openpyxl
import matplotlib.pyplot as plt


mu = 103.94  #期望为1
sigma = 6.64  #标准差为3
num = 10  #个数为10000

rand_data = np.random.normal(mu, sigma, num)
print(rand_data)
count, bins, ignored = plt.hist(rand_data, 30, normed=True)
plt.plot(bins, 1/(sigma * np.sqrt(2 * np.pi)) *np.exp( - (bins - mu)**2 / (2 * sigma**2)), linewidth=2, color='r')
plt.show()

# workbook = openpyxl.Workbook()
# sheet1 = workbook.create_sheet('vehicle',0)
# sheet2 = workbook.create_sheet('obj',1)
#
#
# for i in range(500):
#     speed1 = np.random.randint(low=61, high=120)
#     speed2 = np.random.randint(low=60, high=speed1)
#
#     sheet1.cell(row=i+3, column=1).value = 1
#     sheet1.cell(row=i+3, column=2).value = 'YL3_'+str(i)
#     sheet1.cell(row=i + 3, column=3).value = '小型汽车'
#
#     sheet1.cell(row=i + 3, column=5).value = speed1
#     sheet1.cell(row=i + 3, column=6).value = 0
#
#     sheet1.cell(row=i + 3, column=8).value = 0
#
#     sheet1.cell(row=i + 3, column=9).value = '循线行驶'
#
#
#
#     sheet2.cell(row=i + 3, column=1).value = 'YL3_'+str(i)
#     sheet2.cell(row=i + 3, column=2).value = 2
#     sheet2.cell(row=i + 3, column=3).value = '小型汽车'
#
#     sheet2.cell(row=i + 3, column=4).value = speed2
#     sheet2.cell(row=i + 3, column=5).value = 0
#
#     sheet2.cell(row=i + 3, column=6).value = np.random.randint(0, 500)
#     sheet2.cell(row=i + 3, column=7).value = np.random.rand()
#
#     sheet2.cell(row=i + 3, column=8).value = 0
#
#     sheet2.cell(row=i + 3, column=9).value = '循线行驶'
#
# workbook.save('data/test0204.xlsx')
# workbook.close()