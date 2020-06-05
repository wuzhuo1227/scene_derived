from openpyxl import load_workbook

import pandas as pd
from scipy import stats
import numpy as np

import matplotlib.pyplot as plt
from sklearn.linear_model import LinearRegression

df_object = pd.read_csv('data/nds-sync-object1.csv',encoding="ISO-8859-1")
df_vehicle = pd.read_csv('data/nds-sync-vehicle1.csv',encoding="ISO-8859-1")

# print(df_vehicle.index())
# print(df_vehicle.columns())
# print(df_vehicle)

# print(df_vehicle.dtypes)
# print(df_object.dtypes)
# print(df_vehicle['Time'])

# w_object = load_workbook(u'data/nds-sync-object-14.csv')
# sheet_object = w_object.active
# w_vehicle = load_workbook(u'data/nds-sync-vehicle-14.csv')
# sheet_vehicle = w_vehicle.active


workbook = load_workbook(u'data/ScenariosLabeling2tianda.xlsx')    #找到需要xlsx文件的位置
booksheet = workbook.active                 #获取当前活跃的sheet,默认是第一个sheet

#获取sheet页的行数据
rows = booksheet.rows
#获取sheet页的列数据
columns = booksheet.columns

# print(booksheet["H"])

AdditionalDescription = booksheet['H']
OPosition = booksheet['L']
CBehavior = booksheet['G']
speed = []
id_speed = []
speed_diff = []

a = []
id_a = []

speed = df_vehicle.loc[:,'Vehicle Speed[KPH]']
a = df_vehicle.loc[:,'Acceleration-x[m/s2].1']

# speed = df_object.loc[:,'VXAbs']
# a = df_object.loc[:,'AXAbs']

distance = df_object.loc[:,'LC']


# for x in range(len(df_vehicle)) :
#     speed.extend(df_vehicle.loc[x, 'Vehicle Speed[KPH]'])
#     print(speed)
#     a.extend(df_vehicle[x]['Acceleration-x[m/s2].1'].values)
# for x in range(len(AdditionalDescription)):
#     start_time = booksheet.cell(row = x+1, column = 9).value
#     print(start_time)
#
#     speed.extend(df_vehicle[(df_vehicle.Time == start_time + 1)]['Vehicle Speed[KPH]'].values)
#     print(speed)
#     a.extend(df_vehicle[(df_vehicle.Time == start_time + 1)]['Acceleration-x[m/s2].1'].values)


    # if AdditionalDescription[x].value=="变道超车":
    #     # print(booksheet.cell(row = x+1, column = 9).value)
    #     endTime = booksheet.cell(row = x+1, column = 10).value
    #     if OPosition[x].value=="前" and endTime<500:
    #         id = booksheet.cell(row = x+1, column = 16).value
    #         print(id)
    #         print(endTime)
    #         print(df_vehicle[(df_vehicle.Time >= start_time) & (df_vehicle.Time <= endTime)]['Vehicle Speed[KPH]'].values)
    #         speed.extend(df_vehicle[(df_vehicle.Time >= start_time) & (df_vehicle.Time <= endTime)]['Vehicle Speed[KPH]'].values)
    #         print(df_object[(df_object.Time >= start_time) & (df_object.Time <= endTime) & (df_object.PublicID == id)]['LC'].values)
    #         print("------------")
    #         print()
    #     elif OPosition[x].value=="左前":
    #         left_fid = booksheet.cell(row=x + 1, column=16).value
    #     elif OPosition[x].value=="左后":
    #         left_hid = booksheet.cell(row=x + 1, column=16).value
    #     elif OPosition[x].value=="右前":
    #         right_fid = booksheet.cell(row=x + 1, column=16).value
    #     elif OPosition[x].value=="右后":
    #         right_hid = booksheet.cell(row=x + 1, column=16).value

    # print(CBehavior[x].value)

    # if CBehavior[x].value=='变道向左':
    #     if start_time==400:
    #         break
    #     endTime = booksheet.cell(row=x + 1, column=10).value
    #     # if OPosition[x].value=="左前":
    #     #     left_fid = booksheet.cell(row=x + 1, column=16).value
    #     #     print(left_fid)
    #     #     print(start_time)
    #     #     # v_speed = df_vehicle["Vehicle Speed[KPH]"].values.tolist()[x]
    #     #     # o_speed = df_object["VXAbs"].values.tolist()[x]
    #     #     # print(df_vehicle[(df_vehicle.Time >= start_time) & (df_vehicle.Time <= endTime)]['Vehicle Speed[KPH]'].values)
    #     #     # print(df_object[(df_object.Time >= start_time) & (df_object.Time <= endTime) & (df_object.PublicID == left_fid)]['VXAbs'].values)
    #     #     print(df_object[(df_object.Time >= start_time) & (df_object.Time <= endTime) & (df_object.PublicID == left_fid)][
    #     #               'LC'].values)
    #     #     print(df_object[
    #     #               (df_object.Time >= start_time) & (df_object.Time <= endTime) & (df_object.PublicID == left_fid)][
    #     #               'RC'].values)
    #     #     print("------------")
    #     #     print()
    #     #     speed.extend(
    #     #         df_vehicle[(df_vehicle.Time == start_time+1)]['Vehicle Speed[KPH]'].values)
    #     #     id_speed.extend(df_object[(df_object.Time == start_time+1) & (df_object.PublicID == left_fid)]['VXAbs'].values)
    #     #
    #     #     diff = df_vehicle[(df_vehicle.Time == start_time+1)]['Vehicle Speed[KPH]'].values[0] - df_object[(df_object.Time == start_time+1) & (df_object.PublicID == left_fid)]['VXAbs'].values[0]
    #     #     speed_diff.append(diff)
    #     #
    #     #     a.extend(df_vehicle[(df_vehicle.Time == start_time+1)]['Acceleration-x[m/s2].1'].values)
    #     #     id_a.extend(df_object[(df_object.Time == start_time+1) & (df_object.PublicID == left_fid)]['AXAbs'].values)
    #     if OPosition[x].value=="左后":
    #         left_hid = booksheet.cell(row=x + 1, column=16).value
    #         # print(left_hid)
    #         print(left_hid)
    #         print(start_time)
    #         speed.extend(
    #             df_vehicle[(df_vehicle.Time == start_time+1)]['Vehicle Speed[KPH]'].values)
    #         print(df_object[
    #                   (df_object.Time == start_time+1) & (df_object.PublicID == left_hid)][
    #                   'LC'].values)
    #         print(df_object[
    #                   (df_object.Time == start_time+1) & (df_object.PublicID == left_hid)][
    #                   'RC'].values)
    #         print("------------")
    #         print()
    #
    #         id_speed.extend(df_object[(df_object.Time == start_time+1) & (df_object.PublicID == left_hid)]['VXAbs'].values)
    #
    #         diff = df_vehicle[(df_vehicle.Time == start_time+1)]['Vehicle Speed[KPH]'].values[0] - df_object[(df_object.Time == start_time+1) & (df_object.PublicID == left_hid)]['VXAbs'].values[0]
    #         speed_diff.append(diff)
    #
    #         a.extend(df_vehicle[(df_vehicle.Time == start_time+1)]['Acceleration-x[m/s2].1'].values)
    #         id_a.extend(df_object[(df_object.Time == start_time+1) & (df_object.PublicID == left_hid)]['AXAbs'].values)

    # if CBehavior[x].value=='变道向右':
    #     if OPosition[x].value=="右前":
    #         right_fid = booksheet.cell(row=x + 1, column=16).value
    #     elif OPosition[x].value=="右后":
    #         right_hid = booksheet.cell(row=x + 1, column=16).value

# print(speed)
speed = np.array(speed)
# print(speed)
print(speed.shape)
# print(np.max(speed))
# print(np.min(speed))

# print(id_speed)
# id_speed = np.array(id_speed)
# print(id_speed.shape)
# print(np.max(id_speed))
# print(np.min(id_speed))
#
# speed_diff = np.array(speed_diff)
# print(np.max(speed_diff))
# print(np.min(speed_diff))

# print(a)
# a = a.astype('float64')
a = np.array(a)
# print(a)
print(a.shape)
# print(np.max(a))
# print(np.min(a))

# print(distance)
# a = a.astype('float64')
distance = np.array(distance)
# print(distance)
# print(distance.shape)
# print(np.max(distance))
# print(np.min(distance))

# na = []
# for n in range(distance.shape[0]):
#     # print(n)
#     if distance[n] == 'na':
#         # print(n)
#         na.append(n)
#         #speed = np.delete(speed, n, axis=0)
#         #a = np.delete(a, n, axis=0)
#
# print(len(na))



na = []
for n in range(a.shape[0]):
    # print(n)
    if a[n] == 'na':
        # print(n)
        na.append(n)
        #speed = np.delete(speed, n, axis=0)
        #a = np.delete(a, n, axis=0)

print(na)


# print(a.dtype)

# a = a.astype(np.float64)

# print(np.where(np.isnan(a)))
#
speed = np.delete(speed,na, axis=0)
a = np.delete(a,na, axis=0)
print(speed.shape)
print(a.shape)
#
a = a.astype(np.float64)

speed1 = []
a1 = []

for i in range(speed.shape[0]):
    if speed[i] > 110 and speed[i] < 120:
        speed1.append(speed[i])
        a1.append(a[i])


speed1 = [65,75,85,95,105,115]
# a1 = [-2.47,-2.35,-1.89,-1.67,-1.07,-0.677]
a1 = [0.19, 0.97, 1.07, 1.15, 1.1347, 0.737756]

speed1 = np.array(speed1)
a1 = np.array(a1)
print(np.max(a1))
print(np.min(a1))


# print(id_a)
# id_a = np.array(id_a)
# print(id_a.shape)
# print(np.max(id_a))
# print(np.min(id_a))

# print(np.where(np.isnan(a)))

# z1 = np.polyfit(speed, a, 1) # 用1次多项式拟合
# p1 = np.poly1d(z1)
# print(p1) # 在屏幕上打印拟合多项式

a1 = a1.reshape(-1,1)
speed1 = speed1.reshape(-1,1)

reg = LinearRegression().fit(speed1, a1)
print("一元回归方程为:  Y = %.5fX + (%.5f)" % (reg.coef_[0][0], reg.intercept_[0]))
print("R平方为: %s" % reg.score(speed1, a1))

plt.scatter(speed1, a1,  color='black')
plt.plot(speed1, reg.predict(speed1), color='red', linewidth=1)
plt.show()

# print(speed1)
# print(a1)
# tmp = a1
# a1 = speed1
# speed1 = tmp

# coef 为系数，poly_fit 拟合函数
# coef1 = np.polyfit(speed1,a1, 1)
# poly_fit1 = np.poly1d(coef1)
# plt.plot(speed1, poly_fit1(speed1), 'g',label="一阶拟合")
# print(poly_fit1)
#
# coef2 = np.polyfit(speed1,a1, 2)
# poly_fit2 = np.poly1d(coef2)
# plt.plot(speed1, poly_fit2(speed1), 'b',label="二阶拟合")
# print(poly_fit2)

# coef3 = np.polyfit(speed1,a1, 3)
# poly_fit3 = np.poly1d(coef3)
# plt.plot(speed1, poly_fit3(a1), 'y',label="三阶拟合")
# print(poly_fit3)
#
# coef4 = np.polyfit(speed1,a1, 4)
# poly_fit4 = np.poly1d(coef4)
# plt.plot(speed1, poly_fit4(speed1), 'k',label="四阶拟合")
# print(poly_fit4)
#
# coef5 = np.polyfit(speed1,a1, 5)
# poly_fit5 = np.poly1d(coef5)
# plt.plot(speed1, poly_fit5(speed1), 'r:',label="五阶拟合")
# print(poly_fit5)

# plt.scatter(speed1, a1, color='black')
# plt.legend(loc=2)
# plt.show()
