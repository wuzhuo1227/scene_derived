# -*- coding: utf-8 -*-
# 依据相对速度获取纵向距离范围
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from src.core.curve_fit import polynomial_fit

df_object = pd.read_csv('../../data/nds-sync-object-14.csv', encoding="ISO-8859-1")
df_vehicle = pd.read_csv('../../data/nds-sync-vehicle-14.csv', encoding="ISO-8859-1")

df_object_16 = pd.read_csv('../../data/nds-sync-object-16.csv', encoding="ISO-8859-1")
df_vehicle_16 = pd.read_csv('../../data/nds-sync-vehicle-16.csv', encoding="ISO-8859-1")

workbook = load_workbook(u'../../data/ScenariosLabeling2tianda.xlsx')
booksheet = workbook.active

AdditionalDescription = booksheet['H']
OPosition = booksheet['L']
CBehavior = booksheet['G']

speed = []
a = []
obj_speed = []
obj_a = []

lc = []
rc = []

thw = []

# 读取所需原始传感器数据，包括速度、加速度、目标车速度、目标车加速度、纵向距离、横向距离、车头时距
for x in range(207):
    # if (CBehavior[x].value == "变道向左" and OPosition[x].value=="左后") or (CBehavior[x].value == "变道向右" and OPosition[x].value=="右后"):
    if (AdditionalDescription[x].value == "变道超车" and OPosition[x].value == "前"):
        start_time = booksheet.cell(row=x + 1, column=9).value
        endTime = booksheet.cell(row=x + 1, column=10).value
        id = booksheet.cell(row=x + 1, column=16).value

        speed.extend(df_vehicle[(df_vehicle.Time == start_time + 1)]['Vehicle Speed[KPH]'].values)
        a.extend(df_vehicle[(df_vehicle.Time == start_time + 1)]['Acceleration-x[m/s2].1'].values)

        lc.extend(df_object[(df_object.Time == start_time + 1) & (df_object.PublicID == id)]['LC'].values)
        rc.extend(df_object[(df_object.Time == start_time + 1) & (df_object.PublicID == id)]['RC'].values)

        obj_speed.extend(df_object[(df_object.Time == start_time + 1) & (df_object.PublicID == id)]['VXAbs'].values)
        obj_a.extend(df_object[(df_object.Time == start_time + 1) & (df_object.PublicID == id)]['AXAbs'].values)

        thw.extend(df_object[(df_object.Time == start_time + 1) & (df_object.PublicID == id)]['THW'].values)

for x in range(912, 1703, 1):
    # if (CBehavior[x].value == "变道向左" and OPosition[x].value=="左后") or (CBehavior[x].value == "变道向右" and OPosition[x].value=="右后"):
    if (AdditionalDescription[x].value == "变道超车" and OPosition[x].value == "前"):
        start_time = booksheet.cell(row=x + 1, column=9).value
        endTime = booksheet.cell(row=x + 1, column=10).value
        id = booksheet.cell(row=x + 1, column=16).value

        speed.extend(df_vehicle_16[(df_vehicle_16.Time == start_time + 1)]['Vehicle Speed[KPH]'].values)
        a.extend(df_vehicle_16[(df_vehicle_16.Time == start_time + 1)]['Acceleration-x[m/s2].1'].values)

        lc.extend(df_object_16[(df_object_16.Time == start_time + 1) & (df_object_16.PublicID == id)]['LC'].values)
        rc.extend(df_object_16[(df_object_16.Time == start_time + 1) & (df_object_16.PublicID == id)]['RC'].values)

        obj_speed.extend(
            df_object_16[(df_object_16.Time == start_time + 1) & (df_object_16.PublicID == id)]['VXAbs'].values)
        obj_a.extend(
            df_object_16[(df_object_16.Time == start_time + 1) & (df_object_16.PublicID == id)]['AXAbs'].values)

        thw.extend(df_object_16[(df_object_16.Time == start_time + 1) & (df_object_16.PublicID == id)]['THW'].values)
    if len(speed) != len(obj_speed):
        speed.pop()

speed = np.array(speed)
a = np.array(a)
obj_speed = np.array(obj_speed)
obj_a = np.array(obj_a)
# lc = np.array(lc)
# rc = np.array(rc)
thw = np.array(thw)

a = a.astype(np.float64)
speed = speed.astype(np.float64)
obj_speed = obj_speed.astype(np.float64)
obj_a = obj_a.astype(np.float64)

# 删除包含 na 的数据行
del_pos = []
for n in range(speed.shape[0]):
    if thw[n] == 'na':
        del_pos.append(n)

print(del_pos)

speed = np.delete(speed, del_pos, axis=0)
a = np.delete(a, del_pos, axis=0)
obj_speed = np.delete(obj_speed, del_pos, axis=0)
obj_a = np.delete(obj_a, del_pos, axis=0)
lc = np.delete(lc, del_pos, axis=0)
rc = np.delete(rc, del_pos, axis=0)
thw = np.delete(thw, del_pos, axis=0)

# lc = lc.astype(np.float64)
# rc = rc.astype(np.float64)

thw = thw.astype(np.float64)

distance = thw * speed

a_max = []
a_min = []
obj_speed_max = []
obj_speed_min = []
obj_a_max = []
obj_a_min = []
distance_max = []
distance_min = []
lc_max = []
lc_min = []

rc_max = []
rc_min = []

# 将相对速度分区
real_speed1 = [0, 5, 10, 15, 25, 30, 35, 40]

start = -2.5

# index = [0,1,2,3,4,5,6,7,8,9,10,13]

# 获取每个区域内的最大最小值
for i in range(9):
    if distance[np.where((speed - obj_speed > start + i * 5) & (speed - obj_speed < start + i * 5 + 5))].shape[0] == 0:
        print(i)
        continue
    else:
        distance_max.append(
            distance[np.where((speed - obj_speed > start + i * 5) & (speed - obj_speed < start + i * 5 + 5))].max())
        distance_min.append(
            distance[np.where((speed - obj_speed > start + i * 5) & (speed - obj_speed < start + i * 5 + 5))].min())

font = {'family': 'Times New Roman', 'weight': 'normal', 'size': 15}

print(distance_max)
print(distance_min)
distance_max = np.array(distance_max).reshape(-1, 1)
distance_min = np.array(distance_min).reshape(-1, 1)
real_speed1 = np.array(real_speed1).reshape(-1, 1)

real_speed = speed - obj_speed
print(real_speed.max())
print(real_speed.min())

plt.figure(figsize=(20, 10))

plt.subplot(121)
plt.scatter(real_speed, distance, color='black')
plt.ylabel('distance', font)
plt.xlabel('real_speed', font)

degrees = [1, 2, 3, 4]
colors = ['b', 'y', 'g', 'r']

plt.subplot(122)
polynomial_fit(real_speed1, distance_max, degrees, colors, "parameters/real_distance_max", ranges_min=0, ranges_max=40)
polynomial_fit(real_speed1, distance_min, degrees, colors, "parameters/real_distance_min", 0, ranges_min=0, ranges_max=40)

# reg = LinearRegression().fit(real_speed1, distance_max)
# print("一元回归方程为:  Y = %.5fX + (%.5f)" % (reg.coef_[0][0], reg.intercept_[0]))
# print("R平方为: %s" % reg.score(real_speed1, distance_max))
# plt.plot(real_speed1, reg.predict(real_speed1), color='red', linewidth=1)
#
# reg = LinearRegression().fit(real_speed1, distance_min)
# print("一元回归方程为:  Y = %.5fX + (%.5f)" % (reg.coef_[0][0], reg.intercept_[0]))
# print("R平方为: %s" % reg.score(real_speed1, distance_min))
# plt.plot(real_speed1, reg.predict(real_speed1), color='red', linewidth=1)

plt.scatter(real_speed1, distance_max, color='black', marker='x')
plt.scatter(real_speed1, distance_min, color='black')
plt.scatter(real_speed, distance, color='grey')
plt.ylabel('distance', font)
plt.xlabel('real_speed', font)

plt.show()
