from mpl_toolkits import mplot3d
from openpyxl import load_workbook

import pandas as pd
from scipy import stats
import numpy as np

import matplotlib.pyplot as plt
from sklearn.linear_model import LinearRegression

df_object = pd.read_csv('data/nds-sync-object1.csv',encoding="ISO-8859-1")
df_vehicle = pd.read_csv('data/nds-sync-vehicle1.csv',encoding="ISO-8859-1")

df_object_16 = pd.read_csv('data/nds-sync-object16.csv',encoding="ISO-8859-1")
df_vehicle_16 = pd.read_csv('data/nds-sync-vehicle16.csv',encoding="ISO-8859-1")

workbook = load_workbook(u'data/ScenariosLabeling2tianda.xlsx')
booksheet = workbook.active

AdditionalDescription = booksheet['H']
OPosition = booksheet['L']
CBehavior = booksheet['G']

speed = []
a = []
obj_speed =[]
obj_a =[]

lc = []
rc = []

thw = []

for x in range(207):
    if AdditionalDescription[x].value == "变道超车" and OPosition[x].value=="前":
        start_time = booksheet.cell(row=x + 1, column=9).value
        endTime = booksheet.cell(row=x + 1, column=10).value
        id = booksheet.cell(row=x + 1, column=16).value

        speed.extend(df_vehicle[(df_vehicle.Time == start_time + 1)]['Vehicle Speed[KPH]'].values)
        a.extend(df_vehicle[(df_vehicle.Time == start_time + 1)]['Acceleration-x[m/s2].1'].values)

        lc.extend(df_object[(df_object.Time == start_time + 1) & (df_object.PublicID == id)]['LC'].values)
        rc.extend(df_object[(df_object.Time == start_time + 1) & (df_object.PublicID == id)]['RC'].values)

        obj_speed.extend(df_object[(df_object.Time == start_time + 1) & (df_object.PublicID == id)]['VXAbs'].values)
        obj_a.extend(df_object[(df_object.Time == start_time + 1) & (df_object.PublicID == id)]['AXAbs'].values)

        thw.extend(df_object[(df_object.Time == start_time + 1) & (df_object.PublicID == id)]['THW'].values)

for x in range(912, 1703, 1):
    if AdditionalDescription[x].value == "变道超车" and OPosition[x].value=="前":
        start_time = booksheet.cell(row=x + 1, column=9).value
        endTime = booksheet.cell(row=x + 1, column=10).value
        id = booksheet.cell(row=x + 1, column=16).value

        speed.extend(df_vehicle_16[(df_vehicle_16.Time == start_time + 1)]['Vehicle Speed[KPH]'].values)
        a.extend(df_vehicle_16[(df_vehicle_16.Time == start_time + 1)]['Acceleration-x[m/s2].1'].values)

        lc.extend(df_object_16[(df_object_16.Time == start_time + 1) & (df_object_16.PublicID == id)]['LC'].values)
        rc.extend(df_object_16[(df_object_16.Time == start_time + 1) & (df_object_16.PublicID == id)]['RC'].values)

        obj_speed.extend(df_object_16[(df_object_16.Time == start_time + 1) & (df_object_16.PublicID == id)]['VXAbs'].values)
        obj_a.extend(df_object_16[(df_object_16.Time == start_time + 1) & (df_object_16.PublicID == id)]['AXAbs'].values)

        thw.extend(df_object_16[(df_object_16.Time == start_time + 1) & (df_object_16.PublicID == id)]['THW'].values)

speed = np.array(speed)
a = np.array(a)
obj_speed = np.array(obj_speed)
obj_a = np.array(obj_a)
lc = np.array(lc)
rc = np.array(rc)
thw = np.array(thw)

a = a.astype(np.float64)
speed = speed.astype(np.float64)
obj_speed = obj_speed.astype(np.float64)
obj_a = obj_a.astype(np.float64)
# lc = lc.astype(np.float64)
# rc = rc.astype(np.float64)
thw = thw.astype(np.float64)

# distance = thw*speed


del_pos = []
for n in range(speed.shape[0]):
    if speed[n] < 60:
        del_pos.append(n)

print(del_pos)

speed = np.delete(speed,del_pos, axis=0)
a = np.delete(a,del_pos, axis=0)
obj_speed = np.delete(obj_speed,del_pos, axis=0)
obj_a = np.delete(obj_a,del_pos, axis=0)
lc = np.delete(lc,del_pos, axis=0)
rc = np.delete(rc,del_pos, axis=0)
thw = np.delete(thw,del_pos, axis=0)

distance = thw*speed

print(speed)
print(np.max(speed))
print(np.min(speed))

# speed1 = [65,75,85,95,105,115]
# # speed1 = [62.5, 67.5, 72.5, 77.5, 82.5, 87.5, 92.5, 97.5, 102.5, 107.5, 112.5, 117.5]
#
# a_max = []
# a_min = []
# obj_speed_max = []
# obj_speed_min = []
# obj_a_max = []
# obj_a_min = []
# distance_max = []
# distance_min = []
#
# start = 60
#
# for i in range(6):
#     # if a[np.where((speed>start+i*10) & (speed<start+i*10+10))].shape[0] == 0:
#     #     a_max.append(0)
#     #     a_min.append(0)
#     # else:
#     a_max.append(a[np.where((speed>start+i*10) & (speed<start+i*10+10))].max())
#     a_min.append(a[np.where((speed > start + i * 10) & (speed < start + i * 10 + 10))].min())
#
#     obj_speed_max.append(obj_speed[np.where((speed>start+i*10) & (speed<start+i*10+10))].max())
#     obj_speed_min.append(obj_speed[np.where((speed>start+i*10) & (speed<start+i*10+10))].min())
#
#     obj_a_max.append(obj_a[np.where((speed>start+i*10) & (speed<start+i*10+10))].max())
#     obj_a_min.append(obj_a[np.where((speed > start + i * 10) & (speed < start + i * 10 + 10))].min())
#
#     distance_max.append(distance[np.where((speed>start+i*10) & (speed<start+i*10+10))].max())
#     distance_min.append(distance[np.where((speed > start + i * 10) & (speed < start + i * 10 + 10))].min())
#
#
#
# plt.subplot(221)
# plt.scatter(speed1, a_max,  color='black', marker='x')
# plt.scatter(speed1, a_min,  color='black')
# # plt.scatter(speed, a,  color='black')
# plt.ylabel('a')
# plt.xlabel('speed')
#
# plt.subplot(222)
# plt.scatter(speed1, obj_speed_max,  color='black', marker='x')
# plt.scatter(speed1, obj_speed_min,  color='black')
# # plt.scatter(speed, obj_speed,  color='black')
#
# plt.subplot(223)
# plt.scatter(speed1, obj_a_max,  color='black', marker='x')
# plt.scatter(speed1, obj_a_min,  color='black')
# # plt.scatter(speed, obj_a,  color='black')
#
# plt.subplot(224)
# plt.scatter(speed1, distance_max,  color='black', marker='x')
# plt.scatter(speed1, distance_min,  color='black')
# # plt.scatter(speed, lc,  color='black')
#
# plt.show()

distance = np.array(distance)

xdata = speed
ydata = distance
zdata = speed-obj_speed

# print(xdata)
# print(ydata)
# print(zdata.shape)

xdata = np.array(xdata)
ydata = np.array(ydata)
zdata = np.array(zdata)


# xdata = xdata.reshape(-1,1)
# ydata = ydata.reshape(-1,1)
# zdata = zdata.reshape(-1,1)

print(xdata.shape)
print(ydata.shape)
print(zdata.shape)



ax = plt.axes(projection='3d')
ax.scatter3D(xdata, ydata, zdata, cmap='Greens')

plt.ylabel('distance')
plt.xlabel('speed')
# plt.zlabel('rel_speed')

plt.show()